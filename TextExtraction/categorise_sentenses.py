from os import listdir
from os.path import isfile, join
from openpyxl import Workbook
# import string

mypath = '/Users/keleigong/Google Drive/SCRC 2015 work/extract_sentenses_for_NLP_all_keywords'

file_names= [f for f in listdir(mypath) if isfile(join(mypath, f)) and 'DS_Store' not in f]

# keywords_path = '/Users/keleigong/Google Drive/SCRC 2015 work/2014_data/fourth run/d.sentences/CM/'

keywords = [
        "Second tier supplier audit",
        "Second tier supplier enforcement",
        "Ariba spend management",
        "Beroe",
        "Category Management",
        "Long-term category strategy",
        "Category team",
        "Child labor",
        "Continuous supplier improvement process",
        "Sourcing contract management system",
        "Cross-functional category management team",
        "Cross-functional sourcing team",
        "EDI",
        "EICC",
        "Electronic Industry Citizenship Coalition",
        "Enterprise Resource Planning",
        "e-procurement system",
        "ERP",
        "Fair Labor Association",
        "Global sourcing",
        "Green effort",
        "Health Safety Security Environment",
        "Labor Right",
        "Long-term sourcing strategy",
        "Minimum Wage",
        "Supplier terminate",
        "Non-governmental Organization",
        "Oracle",
        "Procurement",
        "Procurement allocation",
        "Procurement team",
        "Procure-to-Pay system",
        "Product life cycle",
        "SAP",
        "Service Level Agreement",
        "Source approved vendor list",
        "Sourcing Process",
        "Sourcing strategy",
        "Spend management",
        "Spend analytics",
        "SRM policy",
        "Strategic sourcing",
        "Supplier  capacity",
        "Supplier allocation",
        "Supplier assessment",
        "Supplier audit",
        "Supplier award",
        "Supplier code of conduct",
        "Supplier collaboration",
        "Supplier continuity planning",
        "Supplier database",
        "Supplier development plan",
        "Supplier diversity program",
        "Supplier enforcement program",
        "Supplier environmental engagement",
        "Supplier evaluation",
        "Supplier expectation",
        "Supplier feedback",
        "Supplier guideline",
        "Supplier lawsuit",
        "Supplier list",
        "Supplier measurements",
        "Supplier meeting",
        "Supplier optimization",
        "Supplier portal",
        "Supplier purchase terms and conditions",
        "Supplier registration",
        "Supplier Relationship Management",
        "Supplier requirement",
        "Supplier risk management",
        "Supplier scorecard",
        "Supplier segmentation",
        "Supplier selection",
        "Supplier tracking",
        "Supplier training",
        "Supplier verification",
        "Supply base capacity",
        "Supply chain management",
        "Supply management system",
        "Supply market analysis",
        "Supply market intelligence",
        "Supply risk analysis",
        "CSR report",
        "Sustainability report",
        "Talent management",
        "Vendor code of conduct",
        "Vendor expectation",
        "Vendor list",
        "Vendor management",
]

categories = [
        "LHR, ES",
        "LHR, ES",
        "SM, SS, CM",
        "CM",
        "CM",
        "CM",
        "CM",
        "LHR",
        "SS, SRM",
        "SM, SS",
        "CM",
        "SS, CM",
        "SM, SS",
        "SM, SS, LHR, ES",
        "SM, SS, LHR, ES",
        "SM, SS, CM",
        "SM, SS",
        "SM, SS, CM",
        "LHR",
        "SRM",
        "ES",
        "SS",
        "LHR",
        "SS",
        "LHR",
        "SM",
        "LHR",
        "SM, SS, CM",
        "SM, SS",
        "SS",
        "SM",
        "SM, SS",
        "ES",
        "SM, SS, CM",
        "SM, SS",
        "SM, SRM",
        "SM, SS",
        "SS",
        "SM, SS, CM",
        "SM, SS, CM",
        "SRM",
        "SS",
        "SS, CM",
        "SM, SS, CM",
        "SM, SS",
        "SM, SS",
        "SM, SS, SRM",
        "SM, SS, SRM, LHR, ES",
        "SS, CM, SRM",
        "SS, SRM",
        "SM",
        "SS, SRM",
        "SM, SS",
        "LHR, ES",
        "ES",
        "SM, SS",
        "SM, SS",
        "CM, SRM",
        "SM, SS",
        "LHR, ES",
        "SM, SRM",
        "SM, SS",
        "SS, SRM",
        "SS, CM",
        "SM, SS, SRM",
        "SM, SS, SRM",
        "SM",
        "SRM",
        "SM, SS, CM, SRM",
        "SS, CM",
        "SM, SS, CM",
        "SS, CM",
        "SM, SS",
        "SM, SS, LHR, ES",
        "SM, SS, LHR, ES",
        "SS",
        "SS, CM",
        "SM, SS, CM, SRM",
        "SM, SS, CM, SRM",
        "CM",
        "CM",
        "SS, CM",
        "LHR, ES",
        "LHR, ES",
        "SRM",
        "SM, SS, SRM, LHR, ES",
        "SM, SS, SRM",
        "SM, SRM",
        "SM, SS, SRM",
]

keywords_by_category = {}

for i in range(len(keywords)):
    keywords_by_category[keywords[i]] = categories[i]

links = {}
sentenses = {}
rows = []

for f in file_names:
    company_name = f.replace('.txt', '')
    links[company_name] = []
    sentenses[company_name] = []
    file_path = join(mypath, f)
    # print(f)
    file = open(file_path, 'r', encoding='ascii', errors='ignore')

    # content = [line.decode('utf-8').strip().replace(u'0xe2', '') for line in file.readlines()]
    lines = file.readlines()
    file.close()
    for line in lines:
        if line[0:4] == 'http':
            links[company_name].append(line)
        elif line[0:4] != '====' and line[0:4] != 'http':
            sentenses[company_name].append(line)

    for sentense_single_line in sentenses[company_name]:
        sentense_list = sentense_single_line.split('!@#$%^')
        sentense_list = [s for s in sentense_list if len(s)>5]
        for sentense in sentense_list:
            contain_keywords = []
            for keyword in keywords:
                if keyword.lower() in sentense.lower():
                    contain_keywords.append(keyword)
            # sentense = str(sentense.encode('utf-8'))[2:-1]
            # sentense = filter(lambda x: x in string.printable, sentense)
            if len(contain_keywords) == 0:
                row = [company_name, 'N/A', 'N/A', str(sentense)]
            else:
                categories_set = []
                for x in contain_keywords:
                    categories_set += keywords_by_category[x].split(',')
                categories_set = list(set([x.strip() for x in categories_set]))
                row = [company_name, '#$#'.join(contain_keywords), ','.join(categories_set), str(sentense)]
            rows.append(row)

f = open('NLP_all_keywords.txt', 'w')

for row in rows:
    print('!@#$%^'.join(row), file=f)

f.close()

wb = Workbook()
previous = rows[0][0]
sheet = wb.create_sheet(0, previous)
sheet_no = 1

for row in rows:
    if row[0] == previous:
        try:
            sheet.append(row)
        except Exception as e:
            print(e)
            print(row)
        previous = row[0]
    else:
        sheet = wb.create_sheet(sheet_no, row[0])
        try:
            sheet.append(row)
        except Exception as e:
            print(e)
            print(row)
        sheet_no += 1
        previous = row[0]

wb.save('NLP.xlsx')

print('done')

# f = open('/Users/keleigong/Google Drive/SCRC 2015 work/extract_sentenses_for_NLP/adt corp.txt', 'r', encoding='utf-8', errors='ignore')
#
# c = f.readlines()




