__author__ = 'keleigong'
from docx import Document
from openpyxl import Workbook
from docx.text.paragraph import Paragraph

from openpyxl import Workbook, load_workbook
class LinkCategory(object):
    def __init__(self, company, link, categories):
        self.company = company
        self.link = link
        if "SRM Evaluation" in categories:
            self.categories = ["SM", "CM", "SS", "SRM"]
        elif "LHR Evaluation" in categories:
            self.categories = ["LHR"]
        elif "Environmental Sustainability Evaluation" in categories:
            self.categories = ["ES"]
        else:
            self.categories = ["SM", "CM", "SS", "SRM", "LHR", "ES"]

    def __eq__(self, other):
        # print(self.link, other.link)
        if (isinstance(other, self.__class__)) and self.link == other.link:
            other.categories += self.categories
            other.categories = list(set(other.categories))
            return True
        return False

    def __ne__(self, other):
        return not self.__eq__(other)


def is_section(paragraph):
    if any(run.underline for run in paragraph.runs):
        return True
    else:
        return False

def is_category(paragraph):
    # print(p.text)
    if len(paragraph.text.strip()) > 2 and paragraph.text.strip()[-1] == ':' and any(run.bold for run in paragraph.runs):
        return True
    elif any(run.bold for run in paragraph.runs) and all(not run.underline for run in paragraph.runs):
        return True
    else:
        return False

def is_brief(paragraph):
    if not is_section(paragraph) and not is_category(paragraph) and paragraph.text[0:4] != 'http':
        return True
    return False

class row:
    def __init__(self, company, section, category, link='N/A', brief='N/A'):
        self.company = company
        self.section = section
        self.category = category
        if link == '':
            self.link = 'N/A'
        else:
            self.link = link

        if brief == '':
            self.brief = 'N/A'
        else:
            self.brief = brief

    def to_list(self):
        return [self.company, self.section, self.category, self.link, self.brief]

    def __str__(self):
        return self.company + '\n' + self.section + '\n' + self.category + '\n' + self.link + '\n' + self.brief

def generate_rows(doc_file):
    doc = Document(doc_file)
    company_name = []
    processed_data = []
    rows = []
    current_company = ''
    current_section = ''
    current_category = ''
    current_link = ''
    current_brief = ''
    # f = open('processed.txt', 'w')

    for idx, p in enumerate(doc.paragraphs):
        if len(p.text.strip()) > 1:
            if 'Heading 1' in p.style.name:
                # print(p.text)
                # company_name.append(p.text)
                previous_p = p
                current_company = p.text.replace('\n', '').strip()
                # print('=========================',file=f)
                # print(current_company)
            else:
                if is_section(p):
                    current_section = p.text
                    previous_p = p
                    # print('----------------', file=f)
                    # print('\t>>>' + str(p.text), file=f)
                elif is_category(p):
                    if not is_section(previous_p):
                        rows.append(row(current_company, current_section, current_category, current_link, current_brief))
                        current_link, current_brief = '', ''
                    current_category = p.text.replace(':', '').replace('N/A', '').replace('NA', '').strip()
                    previous_p = p
                    # print('\t\t>>' + str(p.text.encode('utf8')), file=f)
                elif p.text[0:4] == 'http' or p.text[0:3] == 'www':
                    if current_link != '' and current_brief != '':
                        rows.append(row(current_company, current_section, current_category, current_link, current_brief))
                        current_link, current_brief = '', ''
                    if p.text[0:3] == 'www':
                        current_link = 'http://' + p.text
                    else:
                        current_link = p.text
                    previous_p = p
                    # print('\t\t\t>' + p.text, file=f)
                else:
                    if is_brief(previous_p):
                        current_brief += p.text
                    else:
                        current_brief = p.text
                    previous_p = p
    return rows
            # print('\t\t\t' + str(p.text.encode('utf8')), file=f)
    # if len(p.runs) > 1:
    #     print(idx)
    #     for run in p.runs:
    #         print(run.text + '---')
    #     print('----------')

# f.close()
def output_to_excel(rows, output_file, company_id=None, year=2015):
    '''
    :param rows: corresponding rows to be put into the excel file
    :param output_file: output file name
    :param company_id: mapping from company name to company id
    :param year: which the secondary data belong to
    :return:
    '''
    wb = Workbook()
    sheet = wb.create_sheet('output', 0)
    # sheet.append(['link', 'frequency', 'company', 'rank', 'year', 'query',
    #                   'link_type', 'title', 'domain', 'snippet', 'year'])
    date = "12/31/{}".format(year)
    if company_id:
        no_match = set()
        matched = set()
        sheet.append(['company_id', 'company_name', 'section', 'category', 'link', 'description', 'date'])
        for row in rows:
            data = row.to_list()
            company_name = data[0].upper()
            # if company_id:
            if company_name in company_id:
                data.insert(0, company_id[company_name])
                data.append(date)
                sheet.append(data)
                matched.add(company_name)
            else:
                no_match.add(company_name)
                # else:

                # print("Company name no match", data[0])
        for name in no_match:
            print(name)
        print(len(no_match), len(matched))
    else:
        sheet.append(['company_name', 'section', 'category', 'link', 'description', 'date'])
        for row in rows:
            data = row.to_list()
            data.append(date)
            sheet.append(data)
    wb.save(output_file)
    print('done')


def get_urls_from_docx(filepath):
    '''
    Extract all urls from the secondary data docx file.
    :param filepath:
    :return:
    '''
    rows = generate_rows(filepath)
    res = {}
    for row in rows:
        if row.company in res.keys():
            if len(row.link) > 5:
                link = LinkCategory(row.company, row.link, row.section)
                if link not in res[row.company]:
                    res[row.company].append(link)
        else:
            res[row.company] = []
            if len(row.link) > 5:
                res[row.company].append(LinkCategory(row.company, row.link, row.section))
    # for key, value in res.items():
    #     res[key] = list(set(value))
    return res

def get_urls_from_excel(filepath):
    res = {}
    wb = load_workbook(filepath, read_only=True)
    ws = wb.get_sheet_by_name("Sheet2")
    for row in ws.rows:
        if 'http' in row[5].value:
            if row[0].value not in res.keys():
                res[row[0].value] = [LinkCategory(row[0].value, row[5].value, "ALL")]
            else:
                res[row[0].value].append(LinkCategory(row[0].value, row[5].value, "ALL"))
    return res


def get_company_id(filepath):
    '''
    create a map between company name to company id, the required data is exported from scrc server's database, table
    "query_company".
    :param filepath:
    :return:
    '''
    with open(filepath, 'r') as f:
        companies = f.read().strip().split('\n')
        companies = [x.split(',') for x in companies]
        res = {x[1].upper(): x[0] for x in companies}
        return res


def docx_to_excel(doc_file, output_file, company_id_file=None, year=2015):
    '''
    convert the docx secondary data file to an excel format.
    :param doc_file:
    :param output_file:
    :param company_id_file:
    :param year:
    :return:
    '''
    rows = generate_rows(doc_file)
    if company_id_file:
        company_id = get_company_id(company_id_file)
        output_to_excel(rows, output_file, company_id, year)
    else:
        output_to_excel(rows, output_file, company_id=None, year=year)

if __name__ == "__main__":
    doc_file = "/Users/keleigong/Dropbox/Python/AUTO_Rating/TextExtraction/secondary data/2013 Secondary Data_ORGANIZED.docx"
    output_file = 'secondary data/2013_secondary_data.xlsx'
    company_id_file = "query_company.csv"
    docx_to_excel(doc_file, output_file, company_id_file, 2013)

