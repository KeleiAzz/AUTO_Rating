from openpyxl import Workbook, load_workbook

def assign_id(rating_file, rating_sheet, company_col, id_file):
    wb = load_workbook(rating_file, read_only=True)
    flag = 1
    ws = wb.get_sheet_by_name(rating_sheet)
    with open(id_file, 'r') as f:
        company_id = {}
        for line in f:
            id, company = line.strip().split(",")
            id = int(id)
            company_id[company] = id
    wb2 = Workbook()
    ws2 = wb2.create_sheet("all ratings", 0)
    for row in ws.rows:
        if flag:
            col_names = [cell.value for cell in row]
            company_idx = col_names.index(company_col)
            col_names.insert(0, 'ID')
            ws2.append(col_names)
            flag = 0
        else:
            values = [cell.value for cell in row]
            if values[company_idx] in company_id:
                new_row = [company_id[values[company_idx]]]
            else:
                new_row = ["NA"]
                print("ID for %s doesn't exist" % values[company_idx])
            new_row.extend(values)
            ws2.append(new_row)
    wb2.save(rating_file[0:rating_file.rindex(".")] + "_with_id.xlsx")


if __name__ == "__main__":
    assign_id("Cleaned data.xlsx", "all ratings", "Bloomberg Name", "query_company.csv")