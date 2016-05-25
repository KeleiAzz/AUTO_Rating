from openpyxl import Workbook, load_workbook

filepath = '/home/scrc/program/AUTO_Rating/DataProcessing/Cleaned data_with_id.xlsx'
output_filepath = ''

wb = load_workbook(filepath, read_only=True)

ws = wb['all ratings']

wb2 = Workbook(write_only=True)

ws2 = wb2.create_sheet('ratings', 0)

flag = 1
ws2.append(["date", 'score', 'expire_date', 'category_id', 'company_id'])
for row in ws.rows:
    if flag:
        col_names = [cell.value for cell in row]
        print(col_names)
        if "Year" in col_names:
            year_idx = col_names.index("Year")
        elif "year" in col_names:
            year_idx = col_names.index("year")
        else:
            print("column for year not found")
            break
        category_idx = col_names.index("SM")
        flag = 0
    else:
        for i in range(0,6,1):
            ws2.append(['1/1/' + str(row[year_idx].value), row[category_idx+i].value,
                        '12/31/' + str(row[year_idx].value), i+1, row[0].value])

wb2.save('ratings.xlsx')
