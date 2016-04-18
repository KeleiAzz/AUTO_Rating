__author__ = 'keleigong'
'''
This file is not in use
'''
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
from openpyxl import Workbook


mypath = '/Users/keleigong/Google Drive/SCRC 2015 work/2014_data/fourth run/d.sentences/CM/'


def read_company_names(mypath):
    # mypath = '/Users/keleigong/Google Drive/SCRC 2015 work/2014_data/third run/d.sentences/ES/'
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    if onlyfiles.count('.DS_Store') > 0:
        onlyfiles.pop(onlyfiles.index('.DS_Store'))
    # onlyfiles.pop(onlyfiles.index('ES.csv'))
    onlyfiles = [name.replace('.txt', '').strip() for name in onlyfiles if name.count('txt')]

    # print(onlyfiles)

    wb = load_workbook('/Users/keleigong/Google Drive/SCRC 2015 work/2014_data/third run/human_rating_all.xlsx')
    ws = wb.get_active_sheet()

    out = Workbook(write_only=True)
    rating = out.create_sheet(0, 'rating')

    dir_num = 0
    rating_num = 0
    names = []
    rating.append(['company', 'sm', 'ss', 'cm', 'srm', 'lhr', 'es'])
    count = 0
    for row in ws.rows:
        if any(x in row[0].value for x in onlyfiles):
            rating.append([cell.value for cell in row])
            names.append(row[0].value)
            count += 1
    print(count)
    for i in range(len(onlyfiles)):
        if names.count(onlyfiles[i]) == 0:
            print(onlyfiles[i])

    out.save(str(count) + 'CM.xlsx')

read_company_names(mypath)