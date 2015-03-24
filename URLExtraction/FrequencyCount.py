__author__ = 'keleigong'
# import openpyxl as ox
# process the scrapered urls from google, count each url's frequency.
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook(filename='/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/output_SUS.xlsx', read_only=True)
ws = wb['Sheet1'] # ws is now an IterableWorksheet

company = {}
count = {}
for row in ws.rows:
    tmp = row[0].value.split(' ')[0]
    if tmp in company.keys():
        if row[1].value in company[tmp]:
            count[tmp][company[tmp].index(row[1].value)] += 1
        else:
            company[tmp].append(row[1].value)
            count[tmp].append(1)
    else:
        company[tmp] = []
        count[tmp] = []
count.pop('query')
company.pop('query')

wb2 = Workbook(write_only=True)


frequency_count = {}
for c in company.keys():
    frequency_count[c] = {}
    for i in range(len(count[c])):
        if count[c][i] in frequency_count[c].keys():
            frequency_count[c][count[c][i]].append(company[c][i])
        else:
            frequency_count[c][count[c][i]] = [company[c][i]]

sheet_no = 0
for cp in frequency_count.keys():
    sheet = wb2.create_sheet(sheet_no, cp)
    frequency_list = list(frequency_count[cp].keys())
    frequency_list.sort(reverse=True)
    Row = 2
    for f in frequency_list:
        sheet.append(['URLs with frequency = %d' % f])
        for url in frequency_count[cp][f]:
            sheet.append(['',url])
            Row += 1
    sheet_no += 1
wb2.save('URL_with_frequency_SUS.xlsx')