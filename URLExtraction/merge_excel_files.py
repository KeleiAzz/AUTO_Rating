__author__ = 'keleigong'
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
from openpyxl import Workbook
# from URLExtraction import CategoryAssign


mypath = '/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/50_companies_test/company_urls_processed/'


def merge_excel_files(mypath):
    onlyfiles = [join(mypath, f) for f in listdir(mypath) if f[-4:] == "xlsx" and isfile(join(mypath, f))]
    # print(onlyfiles)
    merged_file = Workbook()
    output = merged_file.create_sheet(0, "output")
    for path in onlyfiles:
        wb = load_workbook(path)
        ws = wb.get_active_sheet()
        for row in ws.rows:
            output.append([cell.value for cell in row])
    merged_file.save(join(mypath, "merged.xlsx"))



merge_excel_files(mypath)