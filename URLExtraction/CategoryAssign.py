__author__ = 'keleigong'

from openpyxl import load_workbook
from openpyxl import Workbook


def assign_category(url_file, category_file):
    keyword_category = get_keyword_category(category_file)
    wb = load_workbook(url_file, read_only=True)
    out = Workbook(write_only=True)
    url_category = out.create_sheet(0, 'url_category')
    ws = wb.get_sheet_by_name('output')

    url_category.append(["company", "url", "frequency", "sm", "ss", "cm", "srm", "lhr", "es",
                         "sm_ratio", "ss_ratio", "cm_ratio", "srm_ratio", "lhr_ratio", "es_ratio",
                         "sm_keywords", "ss_keywords", "cm_keywords",
                        "srm_keywords", "lhr_keywords", 'es_keywords'])
    keyword_total = [45, 56, 28, 24, 17, 15]
    for row in ws.rows:
        if not isinstance(row[1].value, str) and row[1].value > 1:
            queries = row[5].value.split(', ')
            queries = [query.replace(row[2].value + ' ', "") for query in queries]
            # print(queries)
            categories = []
            for query in queries:
                [categories.append(x) for x in keyword_category[query]]
            # categories = '@_@'.join(list(categories))
            sm, sm_keyword = categories.count("SM"), ",".join([query for query in queries if keyword_category[query].count("SM") > 0])
            ss, ss_keyword = categories.count("SS"), ",".join([query for query in queries if keyword_category[query].count("SS") > 0])
            cm, cm_keyword = categories.count("CM"), ",".join([query for query in queries if keyword_category[query].count("CM") > 0])
            srm, srm_keyword = categories.count("SRM"), ",".join([query for query in queries if keyword_category[query].count("SRM") > 0])
            lhr, lhr_keyword = categories.count("LHR"), ",".join([query for query in queries if keyword_category[query].count("LHR") > 0])
            es, es_keyword = categories.count("ES"), ",".join([query for query in queries if keyword_category[query].count("ES") > 0])
            url_category.append([row[2].value, row[0].value, row[1].value, sm, ss, cm, srm, lhr, es,
                                 round(sm/keyword_total[0], 3),
                                 round(ss/keyword_total[1], 3),
                                 round(cm/keyword_total[2], 3),
                                 round(srm/keyword_total[3], 3),
                                 round(lhr/keyword_total[4], 3),
                                 round(es/keyword_total[5], 3),
                                 sm_keyword, ss_keyword, cm_keyword, srm_keyword, lhr_keyword, es_keyword])

    out.save(url_file[:-5] + '_done.xlsx')
    print('done')


def get_keyword_category(category_file):
    wb = load_workbook(category_file, read_only=True)
    ws = wb.get_active_sheet()
    keyword_category = {}
    for row in ws.rows:
        keyword_category[row[0].value.strip()] = [x.strip() for x in row[1].value.split(',')]
    keyword_category.pop('Keywords')
    return keyword_category

assign_category('/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/50_companies_test/company_1-5.xlsx',
                '/Users/keleigong/Dropbox/Python/AUTO_Rating/Final_keywords.xlsx')
