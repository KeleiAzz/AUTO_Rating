__author__ = 'keleigong'
'''
This file is not in use
'''
from openpyxl import load_workbook
from openpyxl import Workbook


def seperate_companies(filename):
    path = '/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/50_companies_test/company_urls_processed/'
    wb = load_workbook(filename)
    ws = wb.get_active_sheet()
    # company_names = list(set(map(lambda x: x.value, ws.column[0])))
    # for company_name in company_names:
    #     out = Workbook(write_only=True)
    #     url_sheet = out.create_sheet(0, 'links')
    urls = {}
    for row in ws.rows:
        if row[0].value in urls.keys():
            urls[row[0].value].append(row[1].value)
        else:
            urls[row[0].value] = []
    urls.pop('company')
    print(urls.keys())
    for company in urls.keys():
        out = Workbook(write_only=True)
        url_sheet = out.create_sheet(0, 'links')
        for url in urls[company]:
            url_sheet.append([url])
        out.save(path + company + '.xlsx')

seperate_companies('/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/50_companies_test/company_urls_processed/merged.xlsx')