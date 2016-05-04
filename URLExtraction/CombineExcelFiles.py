from openpyxl import load_workbook, Workbook
import os

def combine(path, output_file):
    wb_output = Workbook(output_file)
    ws = wb_output.create_sheet("output", 0)
    flag = 1
    col_name = ["keywords", "categories", "count", "link", "title", "num_results_for_query", "company", "link_type",
                "domain", "rank", "snippet"]
    for file in os.listdir(path):
        if ".xlsx" in file and file != output_file:
            file = os.path.join(path, file)
            if flag:
                wb_input = load_workbook(file, read_only=True)
                ws_input = wb_input.get_sheet_by_name('output')
                rows = [row for row in ws_input.rows]
                col_name = [cell.value.strip() for cell in rows[0]]
                print(col_name)
                ws.append(col_name)
                for row in rows[1:]:
                    ws.append([cell.value for cell in row])
                flag = 0
            else:
                wb_input = load_workbook(file, read_only=True)
                ws_input = wb_input.get_sheet_by_name('output')
                rows = [row for row in ws_input.rows]
                names = [cell.value for cell in rows[0]]
                print(names)
                for row in rows[1:]:
                    values = [cell.value for cell in row]
                    dic = {name: value for name, value in zip(names, values)}
                    # print(dic)
                    new_row = []
                    for name in col_name:
                        if name in dic:
                            new_row.append(dic[name])
                        else:
                            new_row.append(None)
                    ws.append(new_row)
    wb_output.save(output_file)


if __name__ == "__main__":
    combine("606", "combined.xlsx")