__author__ = 'keleigong'

# from openpyxl import load_workbook
from openpyxl import Workbook, load_workbook
import operator
import json
import sqlite3
from collections import defaultdict, namedtuple


JSON_FILE = ""
COMPANY_NAME_FILE = ""

category = {'Ariba spend management':	['SM', 'SS', 'CM'],
            'Beroe':	['CM'],
            'California Transparency Supply Chain':	['LHR'],
            'Category Management':	['CM'],
            'Category team':	['CM'],
            'Child labor':	['LHR'],
            'citizenship report':	['LHR', 'ES'],
            'Continuous supplier improvement process':	['SS', 'SRM'],
            'Cross-functional category management team':	['CM'],
            'Cross-functional sourcing team':	['SS', 'CM'],
            'CSR report':	['LHR', 'ES'],
            'EDI':	['SM', 'SS'],
            'EHS':	['LHR'],
            'EICC':	['SM', 'SS', 'LHR', 'ES'],
            'Electronic Industry Citizenship Coalition':	['SM', 'SS', 'LHR', 'ES'],
            'Enterprise Resource Planning':	['SM', 'SS', 'CM'],
            'Environmental Sustainability':	['ES'],
            'e-procurement system':	['SM', 'SS'],
            'ERP':	['SM', 'SS', 'CM'],
            'Fair Labor Association':	['LHR'],
            'Global sourcing':	['SRM'],
            'Green effort':	['ES'],
            'Health Safety Security Environment':	['SS'],
            'Human Rights':	['LHR'],
            'Labor Right':	['LHR'],
            'Long-term category strategy':	['CM'],
            'Long-term sourcing strategy':	['SS'],
            'Minimum Wage':	['LHR'],
            'Non-governmental Organization':	['LHR'],
            'Oracle':	['SM', 'SS', 'CM'],
            'Procurement':	['SM', 'SS'],
            'Procurement allocation':	['SS'],
            'Procurement team':	['SM'],
            'Procure-to-Pay system':	['SM', 'SS'],
            'Product life cycle':	['ES'],
            'Responsible sourcing':	['LHR', 'ES'],
            'Responsible supply chain management':	['LHR', 'ES'],
            'SAP':	['SM', 'SS', 'CM'],
            'Second tier supplier audit':	['LHR', 'ES'],
            'Second tier supplier enforcement':	['LHR', 'ES'],
            'Service Level Agreement':	['SM', 'SS'],
            'Source approved vendor list':	['SM', 'SRM'],
            'Sourcing contract management system':	['SM', 'SS'],
            'Sourcing Process':	['SM', 'SS'],
            'Sourcing standards':	['SM', 'SS', 'CM', 'SRM', 'LHR', 'ES'],
            'Sourcing strategy':	['SS'],
            'Spend analytics':	['SM', 'SS', 'CM'],
            'Spend management':	['SM', 'SS', 'CM'],
            'SRM policy':	['SRM'],
            'Strategic sourcing':	['SS'],
            'Supplier':	['SM', 'SS', 'CM', 'SRM'],
            'Supplier  capacity':	['SS', 'CM'],
            'Supplier allocation':	['SM', 'SS', 'CM'],
            'Supplier assessment':	['SM', 'SS'],
            'Supplier audit':	['SM', 'SS'],
            'Supplier award':	['SM', 'SS', 'SRM'],
            'Supplier code of conduct':	['SM', 'SS', 'SRM', 'LHR', 'ES'],
            'Supplier collaboration':	['SS', 'CM', 'SRM'],
            'Supplier continuity planning':	['SS', 'SRM'],
            'Supplier database':	['SM'],
            'Supplier development plan':	['SS', 'SRM'],
            'Supplier diversity program':	['SM', 'SS'],
            'Supplier enforcement program':	['LHR', 'ES'],
            'Supplier environmental engagement':	['ES'],
            'Supplier evaluation':	['SM', 'SS'],
            'Supplier expectation':	['SM', 'SS'],
            'Supplier feedback':	['CM', 'SRM'],
            'Supplier guideline':	['SM', 'SS'],
            'Supplier lawsuit':	['LHR', 'ES'],
            'Supplier list':	['SM', 'SRM'],
            'Supplier management':	['SM', 'SS', 'SRM'],
            'Supplier measurements':	['SM', 'SS'],
            'Supplier meeting':	['SS', 'SRM'],
            'Supplier optimization':	['SS', 'CM'],
            'Supplier portal':	['SM', 'SS', 'SRM'],
            'Supplier purchase terms and conditions':	['SM', 'SS', 'SRM'],
            'Supplier registration':	['SM'],
            'Supplier Relationship Management':	['SRM'],
            'Supplier requirement':	['SM', 'SS', 'CM', 'SRM'],
            'Supplier risk management':	['SS', 'CM'],
            'Supplier scorecard':	['SM', 'SS', 'CM'],
            'Supplier segmentation':	['SS', 'CM'],
            'Supplier selection':	['SM', 'SS'],
            'Supplier summit':	['SS', 'SRM'],
            'Supplier terminate':	['SM'],
            'Supplier tracking':	['SM', 'SS', 'LHR', 'ES'],
            'Supplier training':	['SM', 'SS', 'LHR', 'ES'],
            'Supplier verification':	['SS'],
            'Supply base capacity':	['SS', 'CM'],
            'Supply chain management':	['SM', 'SS', 'CM', 'SRM'],
            'Supply management system':	['SM', 'SS', 'CM', 'SRM'],
            'Supply market analysis':	['CM'],
            'Supply market intelligence':	['CM'],
            'Supply risk analysis':	['SS', 'CM'],
            'Sustainability report':	['LHR', 'ES'],
            'Talent management':	['SRM'],
            'Vendor code of conduct':	['SM', 'SS', 'SRM', 'LHR', 'ES'],
            'Vendor expectation':	['SM', 'SS', 'SRM'],
            'Vendor list':	['SM', 'SRM'],
            'Vendor management':	['SM', 'SS', 'SRM'],
            'Vendor portal':	['SM', 'SS', 'CM', 'SRM'],}

class SearchQuery(object):
    def __init__(self, company, query):
        self.company = company
        # self.keyword = ' '.join(query['query'].split(' ')[len(company.split(' ')):])
        self.keyword = query['query'].replace(company, '').strip()
        # print(self.keyword, '---', self.company, query['query'])
        # self.id = query['id']
        self.num_results_for_query = query["num_results_for_query"].replace(",", "")
        self.num_results_for_query = [x for x in self.num_results_for_query.split() if x.isdigit()][0]
        try:
            self.num_results_for_query = int(self.num_results_for_query)
        except Exception as e:
            print(e)
        self.filter = ['erp', 'ariba', 'oracle', 'sap', 'sage', 'edi', 'supplier', 'vendor', 'supply']
        self.results = []
        for result in query['results']:
            if result['link_type'] == 'results' and result['link'][0:4] == 'http':
                self.results.append(SearchResult(company, result, self.keyword, self.num_results_for_query))

class SearchResult(object):
    def __init__(self, company, result, keyword, num_results):
        self.company = company
        self.domain = result['domain']
        # self.id = result['id']
        self.link = result['link']
        self.link_type = result['link_type']
        self.rank = result['rank']
        # self.serp_id = result['serp_id']
        self.snippet = result['snippet']
        self.title = result['title']
        self.visible_link = result['visible_link']
        self.count = 1
        self.keywords = [keyword]
        self.categories = []
        self.num_results_for_query = num_results
        for keyword in self.keywords:
            self.categories += category[keyword.strip()]
        self.categories = list(set(self.categories))

    def set_link(self, link):
        self.link = link

    def update_categories(self):
        self.categories = []
        for keyword in self.keywords:
            self.categories += category[keyword.strip()]
        self.categories = list(set(self.categories))

    def __eq__(self, other):
        # print(self.link, other.link)
        if (isinstance(other, self.__class__)) and self.link == other.link:
            other.count += 1
            other.rank = int(self.rank) + int(other.rank)
            other.num_results_for_query = int(self.num_results_for_query) + int(other.num_results_for_query)
            other.keywords.append(self.keywords[0])
            return True
        return False

    def __ne__(self, other):
        return not self.__eq__(other)


def get_company_names(file_path):
    with open(file_path, 'r') as file:
        companies_names = file.read()
    companies_names = companies_names.strip().split('\n')
    return companies_names


def get_company_query_from_json(json_file, company_file):
    '''
    return a dict, keys are company names, values are queries belong to that company.
    :param json_file:
    :param company_file:
    :return:
    '''
    with open(json_file, 'r') as file:
        json_str = file.read()
        json_data = json.loads(json_str)
    company_names = get_company_names(company_file)
    company_json = {}
    for name in company_names:
        company_json[name] = []
    for query_result in json_data:
        for name in company_names:
            if name in query_result['query']:
                company_json[name].append(query_result)
                break
    return company_json


def query_processing(company_json):
    company_querys = {}
    # print(company_json)
    for company in company_json.keys():
        company_querys[company] = []
    for company, querys in company_json.items():
        for query in querys:
            company_querys[company].append(SearchQuery(company, query))
    return company_querys


def get_company_query_from_db(company_file, db_file=None):
    '''
    read query data from .db file, sometimes the json file may get corrupted.
    :param company_file:
    :return: dict, keys are company names, values are their query results.
    {'company_1': [{'query': "query 1", 'results': [....]},
                   {'query': "query 2", 'results': [....]}],
     'company_2': [{'query': "query 1", 'results': [....]},
                   {'query': "query 2", 'results': [....]}]
    }
    '''
    if not db_file:
        db_file = company_file + '.db'
    def dict_factory(cursor, row):
        d = {}
        for idx, col in enumerate(cursor.description):
            d[col[0]] = row[idx]
        return d
    conn = sqlite3.connect(db_file)
    conn.row_factory = dict_factory
    c = conn.cursor()
    company_json = defaultdict(dict)
    with open(company_file, 'r') as f:
        sql = "select DISTINCT title, snippet, link, visible_link, domain, rank, link_type, query, " \
              "num_results_for_query from link as l, serp as s where l.serp_id=s.id"
        names = f.read().strip().split('\n')
        for row in c.execute(sql):
            for name in names:
                name = name.strip()
                if name in row['query']:
                    tmp = row.copy()
                    tmp.pop('query')
                    tmp.pop('num_results_for_query')
                    if row['query'] in company_json[name]:
                        company_json[name][row['query']]['results'].append(tmp)
                    else:
                        company_json[name][row['query']] = {'query': row['query'],
                                                            'num_results_for_query': row['num_results_for_query'],
                                                            # 'id': row['id'],
                                                            'results':[tmp]}
                    break
        for name in company_json:
            company_json[name] = company_json[name].values()
    return company_json


def remove_irrelevant_urls(company_querys):
    keywords_to_delete = ['linkedin', 'linkup', 'disabledperson', 'indeed', 'simplyhired',
                          'career', 'recruit', 'glassdoor', 'jobs', 'monster.com', 'yelp', 'itunes',
                          'googleadservices', 'wiki', 'www.google.com', 'amazon.com']
    company_all_urls = {}
    for company, querys in company_querys.items():
        company_all_urls[company] = []
        for query in querys:
            for result in query.results:
                if any(k in result.domain for k in keywords_to_delete):
                    continue
                if result in company_all_urls[company]:
                    pass
                else:
                    company_all_urls[company].append(result)
    for company, urls in company_all_urls.items():
        for url in urls:
            url.update_categories()
        urls.sort(key=lambda x: x.count,reverse=True)
        company_all_urls[company] = list(filter(lambda x: x.count>2, urls))
    return company_all_urls

def write_to_xlsx(company_all_urls, filename):
    wb2 = Workbook(filename)
    sheet = wb2.create_sheet('output', 0)
    flag = 1
    print(len(company_all_urls.keys()))
    for company, urls in company_all_urls.items():
        if len(urls) == 0:
            continue
        if flag == 1:
            sheet.append(list(urls[0].__dict__.keys()))
            flag = 0
        for url in urls:
            url.keywords = ', '.join(url.keywords)
            url.categories = ','.join(url.categories)
            row = [value for key, value in url.__dict__.items()]
            sheet.append(row)

    wb2.save(filename)
    return filename





if __name__ == "__main__":
    # file_names = ["1-30", "31-60", "61-90", "91-120", "121-150", "151-180", "181-210",
    #               "211-240", "241-270", '271-300', '301-330', '331-360', '361-390', '391-420',
    #               '421-450', '451-480', '481-520', '521-563', '564-606']
    # for file_name in file_names:
    #     company_json = json_processing('/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/606/%s.json' % (file_name, ),
    #                            '/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/606/%s' % (file_name, ))
    #     company_querys = query_processing(company_json)
    #
    #     company_all_urls = remove_irrelevant_urls(company_querys)
    #
    #     write_to_xlsx(company_all_urls, "/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/606/%s.xlsx" % (file_name,))
    company_json = get_company_query_from_db("606/1-30")
    company_querys = query_processing(company_json)
    company_all_urls = remove_irrelevant_urls(company_querys)
    write_to_xlsx(company_all_urls, "/Users/keleigong/Dropbox/Python/AUTO_Rating/URLExtraction/606/1-30_db.xlsx")

